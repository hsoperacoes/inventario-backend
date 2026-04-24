from fastapi import FastAPI, HTTPException, Depends, UploadFile, File, Query, Request
from fastapi.responses import Response
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from sqlalchemy.orm import Session
from sqlalchemy import func, Column, Integer, String, DateTime, text
from database import Base, engine, SessionLocal
from models import Inventario, Grupo, UsuarioAtivo, Bipe, Estoque
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import csv
import io
import re
from collections import defaultdict
from datetime import datetime, timedelta
from urllib.parse import quote
from typing import Optional


def agora_brasil():
    return agora_brasil() - timedelta(hours=3)


def _import_reportlab():
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.pdfgen import canvas
    from reportlab.graphics.barcode import createBarcodeDrawing
    from reportlab.graphics import renderPDF
    return A4, mm, canvas, createBarcodeDrawing, renderPDF


app = FastAPI(title="HS Inventário API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

Base.metadata.create_all(bind=engine)

MASTER_PASSWORD = "hs1001"
USER_LOCK_NOTICES = {}


def require_admin(request: Request):
    senha = request.headers.get("x-admin-pass", "")
    if senha != MASTER_PASSWORD:
        raise HTTPException(status_code=403, detail="Acesso negado")


def set_user_lock_notice(id_inventario, usuario, grupo, quem_fechou, acao="CONCLUIDO"):
    USER_LOCK_NOTICES[(norm_txt(id_inventario), norm_txt(usuario))] = {
        "status": "SECAO_BLOQUEADA",
        "quemFechou": norm_txt(quem_fechou),
        "grupo": norm_txt(grupo),
        "acao": norm_txt(acao or "CONCLUIDO"),
    }


def clear_user_lock_notice(id_inventario, usuario):
    USER_LOCK_NOTICES.pop((norm_txt(id_inventario), norm_txt(usuario)), None)


def get_user_lock_notice(id_inventario, usuario):
    return USER_LOCK_NOTICES.get((norm_txt(id_inventario), norm_txt(usuario)))


def clear_inventory_lock_notices(id_inventario):
    alvo = norm_txt(id_inventario)
    for k in list(USER_LOCK_NOTICES.keys()):
        if k[0] == alvo:
            USER_LOCK_NOTICES.pop(k, None)


# ── ETIQUETAS PENDENTES (modelo inline) ─────────────────────────────────────
class EtiquetaPendente(Base):
    __tablename__ = "etiquetas_pendentes"

    id = Column(Integer, primary_key=True, autoincrement=True)
    ean = Column(String(32), nullable=False, index=True)
    ref_cor = Column(String(255), nullable=False, default="")
    grade = Column(String(64), nullable=False, default="")
    id_inventario = Column(String(64), nullable=False, index=True)
    id_grupo = Column(String(64), nullable=False, default="")
    usuario = Column(String(120), nullable=False, default="")
    criado_em = Column(DateTime, nullable=False, default=agora_brasil)


Base.metadata.create_all(bind=engine)


# ── PYDANTIC MODELS ──────────────────────────────────────────────────────────
class InventarioIn(BaseModel):
    id: str
    nome: str
    senha: str


class GrupoIn(BaseModel):
    id: str
    id_inventario: str
    nome: str
    meta: int
    colaborativo: bool = False
    vagas: int = 1


class EntrarGrupoIn(BaseModel):
    usuario: str
    id_inventario: str
    id_grupo: str


class BipeIn(BaseModel):
    usuario: str
    id_inventario: str
    id_grupo: str
    ean: str


class ManualBipeIn(BaseModel):
    usuario: str
    id_inventario: str
    id_grupo: str
    ean: str
    secao: Optional[str] = "MANUAL"


class ConcluirGrupoIn(BaseModel):
    usuario: str
    id_inventario: str
    id_grupo: str


class ResetarGrupoIn(BaseModel):
    usuario: str
    id_inventario: str
    id_grupo: str


class ZerarContagemIn(BaseModel):
    usuario: str
    id_inventario: str
    id_grupo: str


class EditarMetaIn(BaseModel):
    id_inventario: str
    id_grupo: str
    nova_meta: int


class TornarColaborativoIn(BaseModel):
    id_inventario: str
    id_grupo: str
    vagas: int = 2


class RenomearGrupoIn(BaseModel):
    id_inventario: str
    id_grupo: str
    novo_nome: str


class RemoverDoGrupoIn(BaseModel):
    usuario: str
    id_inventario: str


class ConsolidadoUpdateIn(BaseModel):
    ean: Optional[str] = None


# ── HELPERS ──────────────────────────────────────────────────────────────────
def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


def normalizar_ean(valor):
    s = str(valor or "").strip()
    s = re.sub(r"\.0+$", "", s)
    s = re.sub(r"[^0-9]", "", s)
    return s


def norm_txt(valor):
    return str(valor or "").strip()


def montar_ref_cor(produto, cor):
    return f"{norm_txt(produto)}{norm_txt(cor)}".strip()


def obter_grupo_ativo_do_usuario(db, usuario, id_inventario):
    return db.query(UsuarioAtivo).filter(
        UsuarioAtivo.usuario == usuario,
        UsuarioAtivo.id_inventario == id_inventario
    ).first()


def contar_bipes_grupo(db, id_inventario, id_grupo):
    """Retorna a soma de quantidade (bipes agrupados)."""
    result = db.query(func.coalesce(func.sum(Bipe.quantidade), 0)).filter(
        Bipe.id_inventario == id_inventario,
        Bipe.id_grupo == id_grupo
    ).scalar()
    return int(result or 0)


def listar_membros_grupo(db, id_inventario, id_grupo):
    return db.query(UsuarioAtivo).filter(
        UsuarioAtivo.id_inventario == id_inventario,
        UsuarioAtivo.id_grupo == id_grupo
    ).all()


def buscar_item_estoque_por_ean(db, ean):
    ean_norm = normalizar_ean(ean)
    if not ean_norm:
        return None
    candidatos = db.query(Estoque).filter(Estoque.ativo.is_(True)).all()
    for item in candidatos:
        if normalizar_ean(getattr(item, "ean", "")) == ean_norm:
            return item
    return None


def registrar_etiqueta_manual(db, *, ean, id_inventario, id_grupo, usuario):
    item = buscar_item_estoque_por_ean(db, ean)
    if not item:
        return None
    ref_cor = norm_txt(getattr(item, "ref_cor", "")) or montar_ref_cor(
        getattr(item, "produto", ""), getattr(item, "cor_produ", "")
    )
    grade = norm_txt(getattr(item, "tamanho", ""))
    etiqueta = EtiquetaPendente(
        ean=normalizar_ean(ean),
        ref_cor=ref_cor,
        grade=grade,
        id_inventario=norm_txt(id_inventario),
        id_grupo=norm_txt(id_grupo),
        usuario=norm_txt(usuario),
    )
    db.add(etiqueta)
    db.commit()
    db.refresh(etiqueta)
    return etiqueta


def listar_etiquetas_pendentes(db, id_inventario=""):
    query = db.query(EtiquetaPendente).order_by(EtiquetaPendente.id.asc())
    if norm_txt(id_inventario):
        query = query.filter(EtiquetaPendente.id_inventario == norm_txt(id_inventario))
    rows = query.all()
    return [
        {
            "id": row.id,
            "ean": norm_txt(row.ean),
            "refCor": norm_txt(row.ref_cor),
            "grade": norm_txt(row.grade),
            "data": row.criado_em.isoformat() if row.criado_em else "",
            "idInventario": norm_txt(row.id_inventario),
            "idGrupo": norm_txt(row.id_grupo),
            "usuario": norm_txt(row.usuario),
        }
        for row in rows
    ]


def gerar_pdf_etiquetas_bytes(etiquetas):
    try:
        A4, mm, canvas, createBarcodeDrawing, renderPDF = _import_reportlab()
    except Exception as e:
        raise RuntimeError("Instale reportlab: pip install reportlab") from e

    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    page_w, page_h = A4
    cols, rows_pg = 4, 4
    margin_x, margin_y = 8 * mm, 8 * mm
    gap_x, gap_y = 4 * mm, 4 * mm
    label_w = (page_w - (2 * margin_x) - ((cols - 1) * gap_x)) / cols
    label_h = (page_h - (2 * margin_y) - ((rows_pg - 1) * gap_y)) / rows_pg

    def draw_label(x, y, e):
        c.setLineWidth(0.8)
        c.rect(x, y, label_w, label_h)
        grade_w, grade_h = 14 * mm, 14 * mm
        c.rect(x + 3 * mm, y + label_h - grade_h - 3 * mm, grade_w, grade_h)
        c.setFont('Helvetica-Bold', 10)
        c.drawCentredString(x + 3 * mm + grade_w / 2, y + label_h - 10 * mm, norm_txt(e.get('grade')) or '?')
        ref_cor = norm_txt(e.get('refCor')) or '—'
        c.setFont('Helvetica-Bold', 6.8)
        text = c.beginText(x + 3 * mm + grade_w + 3 * mm, y + label_h - 7 * mm)
        max_chars = 18
        for part in [ref_cor[i:i + max_chars] for i in range(0, min(len(ref_cor), max_chars * 2), max_chars)]:
            text.textLine(part)
        c.drawText(text)
        ean = normalizar_ean(e.get('ean'))
        if len(ean) == 13:
            try:
                barcode = createBarcodeDrawing('EAN13', value=ean, humanReadable=False, barHeight=12 * mm, width=label_w - 10 * mm)
                renderPDF.draw(barcode, c, x + 5 * mm, y + 16 * mm)
            except Exception:
                c.setFont('Helvetica', 8)
                c.drawString(x + 5 * mm, y + 24 * mm, ean)
        else:
            c.setFont('Helvetica', 8)
            c.drawString(x + 5 * mm, y + 24 * mm, ean)
        c.setFont('Helvetica', 7.5)
        c.drawCentredString(x + label_w / 2, y + 8 * mm, ean)

    for idx, e in enumerate(etiquetas):
        page_pos = idx % (cols * rows_pg)
        col = page_pos % cols
        row = page_pos // cols
        x = margin_x + col * (label_w + gap_x)
        y = page_h - margin_y - (row + 1) * label_h - row * gap_y
        draw_label(x, y, e)
        if page_pos == (cols * rows_pg) - 1 and idx != len(etiquetas) - 1:
            c.showPage()

    if not etiquetas:
        c.setFont('Helvetica', 12)
        c.drawString(20 * mm, page_h - 20 * mm, 'Nenhuma etiqueta pendente.')

    c.save()
    buffer.seek(0)
    return buffer.getvalue()


def montar_label_estoque(produto, cor, tamanho):
    partes = [norm_txt(produto), norm_txt(cor), norm_txt(tamanho)]
    return " ".join([p for p in partes if p]).strip()


def agregar_estoque_por_ean(db):
    itens = db.query(Estoque).filter(Estoque.ativo.is_(True)).all()
    estoque = {}
    total_estoque = 0
    for item in itens:
        ean = normalizar_ean(item.ean)
        if not ean:
            continue
        qtd = int(item.quantidade or 0)
        total_estoque += qtd
        if ean not in estoque:
            estoque[ean] = {
                "ean": ean,
                "qtdEstoque": 0,
                "label": montar_label_estoque(item.produto, item.cor_produ, item.tamanho),
                "ref": norm_txt(item.produto),
                "cor": norm_txt(item.cor_produ),
                "grade": norm_txt(item.tamanho),
                "tamanho": norm_txt(item.tamanho),
                "refCor": norm_txt(item.ref_cor),
            }
        estoque[ean]["qtdEstoque"] += qtd
        if not estoque[ean]["label"]:
            estoque[ean]["label"] = montar_label_estoque(item.produto, item.cor_produ, item.tamanho)
    return estoque, total_estoque


def agregar_bipes_por_ean(db, id_inventario):
    """
    Retorna {ean: total_quantidade} apenas de grupos CONCLUIDOS.
    Como bipes são agrupados, soma o campo quantidade.
    """
    grupos_query = db.query(Grupo).filter(Grupo.status == "CONCLUIDO")
    if id_inventario:
        grupos_query = grupos_query.filter(Grupo.id_inventario == id_inventario)
    grupos_concluidos_ids = {str(g.id) for g in grupos_query.all()}

    query = db.query(Bipe)
    if id_inventario:
        query = query.filter(Bipe.id_inventario == id_inventario)

    bipados = defaultdict(int)
    total_consolidado = 0
    for row in query.all():
        if str(row.id_grupo) not in grupos_concluidos_ids:
            continue
        ean = normalizar_ean(row.ean)
        if not ean:
            continue
        qtd = int(row.quantidade or 1)
        bipados[ean] += qtd
        total_consolidado += qtd
    return bipados, total_consolidado


def _expandir_bipe_row(b, grupos_map, estoque_por_ean):
    """
    Expande um registro de bipe agrupado em N dicionários (um por unidade),
    mantendo a experiência do painel idêntica ao modelo anterior.
    """
    item = estoque_por_ean.get(normalizar_ean(getattr(b, "ean", "")))
    qtd = int(b.quantidade or 1)
    base = {
        "idInventario": norm_txt(b.id_inventario),
        "idGrupo": norm_txt(b.id_grupo),
        "grupo": norm_txt(b.grupo_nome) or norm_txt(
            getattr(grupos_map.get(str(b.id_grupo)), "nome", "")
        ),
        "usuario": norm_txt(b.usuario),
        "ean": normalizar_ean(getattr(b, "ean", "")),
        "hora": b.atualizado_em.isoformat() if getattr(b, "atualizado_em", None) else (
            b.criado_em.isoformat() if getattr(b, "criado_em", None) else ""
        ),
        "ref": norm_txt(getattr(item, "produto", "")) if item else "",
        "cor": norm_txt(getattr(item, "cor_produ", "")) if item else "",
        "tamanho": norm_txt(getattr(item, "tamanho", "")) if item else "",
        "grade": norm_txt(getattr(item, "tamanho", "")) if item else "",
        "refCor": norm_txt(getattr(item, "ref_cor", "")) if item else "",
        "naoEncontrado": item is None,
    }
    # Gera uma linha por unidade (expande quantidade) para o painel
    rows = []
    for i in range(qtd):
        row = dict(base)
        row["id"] = int(b.id) * 10000 + i   # ID único por linha expandida
        rows.append(row)
    return rows


def listar_consolidado_rows(db, id_inventario=""):
    grupos_query = db.query(Grupo).filter(Grupo.status == "CONCLUIDO")
    if id_inventario:
        grupos_query = grupos_query.filter(Grupo.id_inventario == id_inventario)
    grupos = grupos_query.all()
    grupos_map = {str(g.id): g for g in grupos}
    grupos_ids = set(grupos_map.keys())

    query = db.query(Bipe)
    if id_inventario:
        query = query.filter(Bipe.id_inventario == id_inventario)

    itens_estoque = db.query(Estoque).filter(Estoque.ativo.is_(True)).all()
    estoque_por_ean = {normalizar_ean(getattr(item, "ean", "")): item for item in itens_estoque}

    rows = []
    for b in query.all():
        if str(b.id_grupo) not in grupos_ids:
            continue
        rows.extend(_expandir_bipe_row(b, grupos_map, estoque_por_ean))

    rows.sort(key=lambda x: (x["idInventario"], x["grupo"], x["usuario"], x["id"]))
    return rows


def montar_confronto_estoque(db, id_inventario):
    estoque, total_estoque = agregar_estoque_por_ean(db)
    bipados, total_consolidado = agregar_bipes_por_ean(db, id_inventario)

    acima, proximo, abaixo50, exato, nao_encontrados = [], [], [], [], []

    for ean, item_base in estoque.items():
        qtd_bipada = int(bipados.get(ean, 0))
        if qtd_bipada == 0:
            continue
        item = {
            "ean": ean,
            "label": item_base.get("label", ""),
            "ref": item_base.get("ref", ""),
            "cor": item_base.get("cor", ""),
            "grade": item_base.get("grade", ""),
            "tamanho": item_base.get("tamanho", ""),
            "refCor": item_base.get("refCor", ""),
            "qtdEstoque": int(item_base.get("qtdEstoque", 0)),
            "qtdBipada": qtd_bipada,
        }
        if qtd_bipada > item["qtdEstoque"]:
            acima.append(item)
        elif qtd_bipada == item["qtdEstoque"]:
            exato.append(item)
        else:
            pct = 100 if item["qtdEstoque"] == 0 else round((qtd_bipada / item["qtdEstoque"]) * 100)
            item_pct = dict(item)
            item_pct["pct"] = int(pct)
            if pct >= 50:
                proximo.append(item_pct)
            else:
                abaixo50.append(item_pct)

    for ean, qtd_bipada in bipados.items():
        if ean in estoque:
            continue
        nao_encontrados.append({
            "ean": ean, "qtdBipada": int(qtd_bipada), "qtdEstoque": 0,
            "label": "", "ref": "", "cor": "", "grade": "", "tamanho": "", "refCor": "",
        })

    proximo.sort(key=lambda x: (-int(x.get("pct", 0)), str(x.get("label", ""))))
    abaixo50.sort(key=lambda x: (-int(x.get("pct", 0)), str(x.get("label", ""))))
    acima.sort(key=lambda x: -((int(x.get("qtdBipada", 0)) - int(x.get("qtdEstoque", 0)))))
    exato.sort(key=lambda x: str(x.get("label", "")))
    nao_encontrados.sort(key=lambda x: (-int(x.get("qtdBipada", 0)), str(x.get("ean", ""))))

    percentual = round((total_consolidado / total_estoque) * 1000) / 10 if total_estoque > 0 else 0

    return {
        "success": True,
        "acima": acima, "proximo": proximo, "abaixo50": abaixo50,
        "exato": exato, "naoEncontrados": nao_encontrados,
        "totalEstoque": int(total_estoque),
        "totalBipados": len(bipados),
        "totalConsolidado": int(total_consolidado),
        "percentualConsolidado": percentual,
    }


def montar_status_secao(db, id_inventario, usuario):
    inventario = db.get(Inventario, id_inventario)
    if not inventario:
        return {"success": False, "message": "Inventário não encontrado"}

    if norm_txt(inventario.status) != "ABERTO":
        return {"success": True, "status": "INVENTARIO_ENCERRADO", "message": "Inventário encerrado."}

    notice = get_user_lock_notice(id_inventario, usuario)
    if notice:
        return {
            "success": True,
            "status": notice.get("status", "SECAO_BLOQUEADA"),
            "quemFechou": notice.get("quemFechou", ""),
            "grupo": notice.get("grupo", ""),
            "acao": notice.get("acao", "CONCLUIDO"),
        }

    ativo = db.query(UsuarioAtivo).filter(
        UsuarioAtivo.usuario == usuario,
        UsuarioAtivo.id_inventario == id_inventario
    ).first()

    if not ativo:
        return {"success": True, "status": "SEM_GRUPO"}

    grupo = db.query(Grupo).filter(
        Grupo.id == ativo.id_grupo,
        Grupo.id_inventario == id_inventario
    ).first()

    if not grupo:
        return {"success": True, "status": "SEM_GRUPO"}

    total = contar_bipes_grupo(db, id_inventario, ativo.id_grupo)

    membros = db.query(UsuarioAtivo).filter(
        UsuarioAtivo.id_inventario == id_inventario,
        UsuarioAtivo.id_grupo == ativo.id_grupo
    ).all()

    return {
        "success": True,
        "status": "LIBERADO",
        "usuario": usuario,
        "idGrupo": ativo.id_grupo,
        "grupo": ativo.grupo_nome,
        "meta": int(grupo.meta or 0),
        "count": int(total),
        "colaborativo": bool(grupo.colaborativo),
        "membros": [m.usuario for m in membros],
    }


# ── ENDPOINTS ────────────────────────────────────────────────────────────────

@app.get("/")
def home():
    return {"status": "API rodando"}


# ── PING LEVE — usado pelo painel para polling de 1s ─────────────────────────
# Retorna apenas o total de bipes e um hash mínimo de estado.
# O painel só chama /admin/painel quando esse valor mudar.
@app.get("/admin/ping")
def admin_ping(db: Session = Depends(get_db)):
    total_bipes = db.query(func.coalesce(func.sum(Bipe.quantidade), 0)).scalar() or 0
    total_usuarios = db.query(func.count(UsuarioAtivo.id)).scalar() or 0
    return {
        "total_bipes": int(total_bipes),
        "total_usuarios": int(total_usuarios),
        "ts": agora_brasil().isoformat()
    }


@app.get("/status/secao")
def status_secao(usuario: str, id_inventario: str, db: Session = Depends(get_db)):
    return montar_status_secao(db, id_inventario, usuario)


@app.post("/inventarios")
def criar_inventario(data: InventarioIn, db: Session = Depends(get_db)):
    existe = db.get(Inventario, data.id)
    if existe:
        raise HTTPException(status_code=400, detail="Inventário já existe")
    inv = Inventario(id=data.id, nome=data.nome, senha=data.senha, status="ABERTO")
    db.add(inv)
    db.commit()
    return {"success": True}


@app.get("/inventarios")
def listar_inventarios(db: Session = Depends(get_db)):
    itens = db.query(Inventario).all()
    return {
        "success": True,
        "inventarios": [{"id": i.id, "nome": i.nome, "senha": i.senha, "status": i.status} for i in itens]
    }


@app.post("/grupos")
def criar_grupo(data: GrupoIn, db: Session = Depends(get_db)):
    inv = db.get(Inventario, data.id_inventario)
    if not inv:
        raise HTTPException(status_code=404, detail="Inventário não encontrado")
    existe = db.get(Grupo, data.id)
    if existe:
        raise HTTPException(status_code=400, detail="Grupo já existe")
    grupo = Grupo(
        id=data.id, id_inventario=data.id_inventario, nome=data.nome,
        meta=data.meta, status="DISPONIVEL", colaborativo=data.colaborativo, vagas=data.vagas
    )
    db.add(grupo)
    db.commit()
    return {"success": True}


@app.get("/grupos/{id_inventario}")
def listar_grupos(id_inventario: str, db: Session = Depends(get_db)):
    lista = db.query(Grupo).filter(Grupo.id_inventario == id_inventario).all()
    grupos = []
    for g in lista:
        membros = db.query(UsuarioAtivo).filter(
            UsuarioAtivo.id_inventario == g.id_inventario,
            UsuarioAtivo.id_grupo == g.id
        ).all()
        grupos.append({
            "id": g.id, "id_inventario": g.id_inventario, "nome": g.nome,
            "meta": g.meta, "status": g.status, "colaborativo": g.colaborativo,
            "vagas": g.vagas, "membros": [m.usuario for m in membros]
        })
    return {"success": True, "grupos": grupos}


@app.post("/grupos/entrar")
def entrar_grupo(data: EntrarGrupoIn, db: Session = Depends(get_db)):
    grupo = db.query(Grupo).filter(
        Grupo.id == data.id_grupo, Grupo.id_inventario == data.id_inventario
    ).first()
    if not grupo:
        raise HTTPException(status_code=404, detail="Grupo não encontrado")
    if grupo.status == "CONCLUIDO":
        raise HTTPException(status_code=400, detail="Grupo concluído")

    membros = db.query(UsuarioAtivo).filter(
        UsuarioAtivo.id_inventario == data.id_inventario,
        UsuarioAtivo.id_grupo == data.id_grupo
    ).all()

    ja_ativo = db.query(UsuarioAtivo).filter(
        UsuarioAtivo.usuario == data.usuario,
        UsuarioAtivo.id_inventario == data.id_inventario
    ).first()

    if ja_ativo:
        ja_ativo.id_grupo = grupo.id
        ja_ativo.grupo_nome = grupo.nome
    else:
        if not grupo.colaborativo and len(membros) >= 1:
            raise HTTPException(status_code=400, detail="Grupo já reservado")
        if grupo.colaborativo and len(membros) >= grupo.vagas:
            raise HTTPException(status_code=400, detail="Sem vagas")
        novo = UsuarioAtivo(
            usuario=data.usuario, id_inventario=data.id_inventario,
            id_grupo=grupo.id, grupo_nome=grupo.nome
        )
        db.add(novo)

    grupo.status = "RESERVADO"
    clear_user_lock_notice(data.id_inventario, data.usuario)
    db.commit()

    membros_atualizados = db.query(UsuarioAtivo).filter(
        UsuarioAtivo.id_inventario == data.id_inventario,
        UsuarioAtivo.id_grupo == data.id_grupo
    ).all()

    return {
        "success": True, "grupo": grupo.nome, "meta": grupo.meta,
        "colaborativo": grupo.colaborativo,
        "membros": [m.usuario for m in membros_atualizados]
    }


@app.get("/usuario/ativo")
def usuario_ativo(usuario: str, id_inventario: str, db: Session = Depends(get_db)):
    ativo = db.query(UsuarioAtivo).filter(
        UsuarioAtivo.usuario == usuario,
        UsuarioAtivo.id_inventario == id_inventario
    ).first()

    if not ativo:
        aviso = get_user_lock_notice(id_inventario, usuario)
        if aviso:
            return {
                "ativo": False, "status": aviso.get("status", "SECAO_BLOQUEADA"),
                "quemFechou": aviso.get("quemFechou", ""),
                "grupo": aviso.get("grupo", ""), "acao": aviso.get("acao", "CONCLUIDO")
            }
        return {"ativo": False}

    grupo = db.query(Grupo).filter(
        Grupo.id == ativo.id_grupo, Grupo.id_inventario == id_inventario
    ).first()

    if not grupo:
        return {"ativo": False}

    total = contar_bipes_grupo(db, id_inventario, ativo.id_grupo)
    membros = db.query(UsuarioAtivo).filter(
        UsuarioAtivo.id_inventario == id_inventario,
        UsuarioAtivo.id_grupo == ativo.id_grupo
    ).all()

    return {
        "ativo": True, "usuario": usuario,
        "id_grupo": ativo.id_grupo, "grupo_nome": ativo.grupo_nome,
        "meta": grupo.meta, "colaborativo": grupo.colaborativo,
        "bipes": total, "membros": [m.usuario for m in membros]
    }


# ── BIPES: UPSERT AGRUPADO ────────────────────────────────────────────────────
@app.post("/bipes")
def registrar_bipe(data: BipeIn, db: Session = Depends(get_db)):
    grupo = db.query(Grupo).filter(
        Grupo.id == data.id_grupo, Grupo.id_inventario == data.id_inventario
    ).first()
    if not grupo:
        raise HTTPException(status_code=404, detail="Grupo não encontrado")
    if grupo.status == "CONCLUIDO":
        raise HTTPException(status_code=400, detail="Grupo concluído")

    usuario_ok = db.query(UsuarioAtivo).filter(
        UsuarioAtivo.usuario == data.usuario,
        UsuarioAtivo.id_inventario == data.id_inventario,
        UsuarioAtivo.id_grupo == data.id_grupo
    ).first()
    if not usuario_ok:
        raise HTTPException(status_code=400, detail="Usuário não está ativo nesse grupo")

    ean_norm = normalizar_ean(data.ean)

    # Upsert agrupado: incrementa quantidade se já existe, senão insere
    existente = db.query(Bipe).filter(
        Bipe.usuario == data.usuario,
        Bipe.id_inventario == data.id_inventario,
        Bipe.id_grupo == data.id_grupo,
        Bipe.ean == ean_norm
    ).first()

    if existente:
        existente.quantidade += 1
        existente.atualizado_em = agora_brasil()
    else:
        novo = Bipe(
            usuario=data.usuario,
            id_inventario=data.id_inventario,
            id_grupo=data.id_grupo,
            grupo_nome=usuario_ok.grupo_nome,
            ean=ean_norm,
            quantidade=1
        )
        db.add(novo)

    db.commit()
    total_grupo = contar_bipes_grupo(db, data.id_inventario, data.id_grupo)
    return {"success": True, "total_grupo": total_grupo}


@app.post("/bipes/manual")
def registrar_bipe_manual(data: ManualBipeIn, db: Session = Depends(get_db)):
    grupo = db.query(Grupo).filter(
        Grupo.id == data.id_grupo, Grupo.id_inventario == data.id_inventario
    ).first()
    if not grupo:
        raise HTTPException(status_code=404, detail="Grupo não encontrado")

    usuario_ok = db.query(UsuarioAtivo).filter(
        UsuarioAtivo.usuario == data.usuario,
        UsuarioAtivo.id_inventario == data.id_inventario,
        UsuarioAtivo.id_grupo == data.id_grupo
    ).first()
    if not usuario_ok:
        raise HTTPException(status_code=400, detail="Usuário não está ativo nesse grupo")

    ean_norm = normalizar_ean(data.ean)

    existente = db.query(Bipe).filter(
        Bipe.usuario == data.usuario,
        Bipe.id_inventario == data.id_inventario,
        Bipe.id_grupo == data.id_grupo,
        Bipe.ean == ean_norm
    ).first()

    if existente:
        existente.quantidade += 1
        existente.atualizado_em = agora_brasil()
    else:
        novo = Bipe(
            usuario=data.usuario,
            id_inventario=data.id_inventario,
            id_grupo=data.id_grupo,
            grupo_nome=usuario_ok.grupo_nome,
            ean=ean_norm,
            quantidade=1
        )
        db.add(novo)

    db.commit()

    etiqueta = registrar_etiqueta_manual(
        db, ean=data.ean, id_inventario=data.id_inventario,
        id_grupo=data.id_grupo, usuario=data.usuario,
    )

    total_grupo = contar_bipes_grupo(db, data.id_inventario, data.id_grupo)
    return {"success": True, "total_grupo": total_grupo, "etiqueta_gerada": bool(etiqueta)}


@app.post("/grupos/concluir")
def concluir_grupo(data: ConcluirGrupoIn, db: Session = Depends(get_db)):
    grupo = db.query(Grupo).filter(
        Grupo.id == data.id_grupo, Grupo.id_inventario == data.id_inventario
    ).first()
    if not grupo:
        raise HTTPException(status_code=404, detail="Grupo não encontrado")

    usuario_ativo_obj = db.query(UsuarioAtivo).filter(
        UsuarioAtivo.usuario == data.usuario,
        UsuarioAtivo.id_inventario == data.id_inventario,
        UsuarioAtivo.id_grupo == data.id_grupo
    ).first()
    if not usuario_ativo_obj:
        raise HTTPException(status_code=400, detail="Usuário não está ativo nesse grupo")

    total_grupo = contar_bipes_grupo(db, data.id_inventario, data.id_grupo)

    if int(total_grupo) != int(grupo.meta or 0):
        raise HTTPException(status_code=400, detail="CONTAGEM_NAO_BATE")

    membros = db.query(UsuarioAtivo).filter(
        UsuarioAtivo.id_inventario == data.id_inventario,
        UsuarioAtivo.id_grupo == data.id_grupo
    ).all()

    for m in membros:
        set_user_lock_notice(data.id_inventario, m.usuario, grupo.nome, data.usuario, 'CONCLUIDO')

    grupo.status = "CONCLUIDO"
    db.query(UsuarioAtivo).filter(
        UsuarioAtivo.id_inventario == data.id_inventario,
        UsuarioAtivo.id_grupo == data.id_grupo
    ).delete(synchronize_session=False)
    db.commit()

    return {
        "success": True, "grupo": grupo.nome, "id_grupo": grupo.id,
        "count": total_grupo, "membros_removidos": [m.usuario for m in membros],
        "finalizado_por": data.usuario, "forcado": False
    }


@app.post("/grupos/concluir-forcado")
def concluir_grupo_forcado(request: Request, data: ConcluirGrupoIn, db: Session = Depends(get_db)):
    require_admin(request)
    grupo = db.query(Grupo).filter(
        Grupo.id == data.id_grupo, Grupo.id_inventario == data.id_inventario
    ).first()
    if not grupo:
        raise HTTPException(status_code=404, detail="Grupo não encontrado")

    total_grupo = contar_bipes_grupo(db, data.id_inventario, data.id_grupo)
    membros = db.query(UsuarioAtivo).filter(
        UsuarioAtivo.id_inventario == data.id_inventario,
        UsuarioAtivo.id_grupo == data.id_grupo
    ).all()

    for m in membros:
        set_user_lock_notice(data.id_inventario, m.usuario, grupo.nome, data.usuario, 'CONCLUIDO')

    grupo.status = "CONCLUIDO"
    db.query(UsuarioAtivo).filter(
        UsuarioAtivo.id_inventario == data.id_inventario,
        UsuarioAtivo.id_grupo == data.id_grupo
    ).delete(synchronize_session=False)
    db.commit()

    return {
        "success": True, "grupo": grupo.nome, "id_grupo": grupo.id,
        "count": total_grupo, "forcado": True,
        "membros_removidos": [m.usuario for m in membros],
        "finalizado_por": data.usuario
    }


@app.post("/grupos/editar-meta")
def editar_meta(data: EditarMetaIn, db: Session = Depends(get_db)):
    grupo = db.query(Grupo).filter(
        Grupo.id == data.id_grupo, Grupo.id_inventario == data.id_inventario
    ).first()
    if not grupo:
        raise HTTPException(status_code=404, detail="Grupo não encontrado")
    if int(data.nova_meta or 0) <= 0:
        raise HTTPException(status_code=400, detail="Meta inválida")

    grupo.meta = int(data.nova_meta)
    if grupo.status == "CONCLUIDO":
        grupo.status = "RESERVADO" if db.query(UsuarioAtivo).filter(
            UsuarioAtivo.id_inventario == data.id_inventario,
            UsuarioAtivo.id_grupo == data.id_grupo
        ).first() else "DISPONIVEL"
    db.commit()

    total_grupo = contar_bipes_grupo(db, data.id_inventario, data.id_grupo)
    return {"success": True, "grupo": grupo.nome, "id_grupo": grupo.id, "nova_meta": int(grupo.meta or 0), "bipes_atual": int(total_grupo)}


@app.post("/grupos/tornar-colaborativo")
def tornar_colaborativo(data: TornarColaborativoIn, db: Session = Depends(get_db)):
    grupo = db.query(Grupo).filter(
        Grupo.id == data.id_grupo, Grupo.id_inventario == data.id_inventario
    ).first()
    if not grupo:
        raise HTTPException(status_code=404, detail="Grupo não encontrado")

    vagas = int(data.vagas or 2)
    if vagas < 2:
        raise HTTPException(status_code=400, detail="Vagas inválidas")

    membros = listar_membros_grupo(db, data.id_inventario, data.id_grupo)
    grupo.colaborativo = True
    grupo.vagas = vagas
    grupo.status = "RESERVADO" if membros else "DISPONIVEL"
    for m in membros:
        clear_user_lock_notice(data.id_inventario, m.usuario)
    db.commit()

    return {
        "success": True, "grupo": grupo.nome, "id_grupo": grupo.id,
        "colaborativo": True, "vagas": int(grupo.vagas or 0),
        "membros": [m.usuario for m in membros]
    }


@app.post("/grupos/renomear")
def renomear_grupo(data: RenomearGrupoIn, db: Session = Depends(get_db)):
    grupo = db.query(Grupo).filter(
        Grupo.id == data.id_grupo, Grupo.id_inventario == data.id_inventario
    ).first()
    if not grupo:
        raise HTTPException(status_code=404, detail="Grupo não encontrado")

    novo_nome = norm_txt(data.novo_nome).upper()
    if not novo_nome:
        raise HTTPException(status_code=400, detail="Novo nome inválido")

    grupo.nome = novo_nome
    db.query(UsuarioAtivo).filter(
        UsuarioAtivo.id_inventario == data.id_inventario,
        UsuarioAtivo.id_grupo == data.id_grupo
    ).update({UsuarioAtivo.grupo_nome: novo_nome}, synchronize_session=False)
    db.query(Bipe).filter(
        Bipe.id_inventario == data.id_inventario,
        Bipe.id_grupo == data.id_grupo
    ).update({Bipe.grupo_nome: novo_nome}, synchronize_session=False)
    db.commit()

    return {"success": True, "grupo": novo_nome, "id_grupo": grupo.id}


@app.post("/grupos/remover-do-grupo")
def remover_do_grupo(data: RemoverDoGrupoIn, db: Session = Depends(get_db)):
    ativo = db.query(UsuarioAtivo).filter(
        UsuarioAtivo.usuario == data.usuario,
        UsuarioAtivo.id_inventario == data.id_inventario
    ).first()
    if not ativo:
        raise HTTPException(status_code=404, detail="Usuário não está em grupo ativo")

    grupo = db.query(Grupo).filter(
        Grupo.id == ativo.id_grupo, Grupo.id_inventario == data.id_inventario
    ).first()

    nome_grupo = ativo.grupo_nome
    id_grupo = ativo.id_grupo
    set_user_lock_notice(data.id_inventario, data.usuario, nome_grupo, "Administrador", "REMOVIDO")
    db.delete(ativo)
    db.flush()

    restantes = listar_membros_grupo(db, data.id_inventario, id_grupo)
    if grupo:
        grupo.status = "RESERVADO" if restantes else "DISPONIVEL"

    total_grupo = contar_bipes_grupo(db, data.id_inventario, id_grupo)
    db.commit()

    return {
        "success": True, "usuario_removido": data.usuario,
        "grupo": nome_grupo, "id_grupo": id_grupo,
        "membros_restantes": [m.usuario for m in restantes],
        "bipes_mantidos": int(total_grupo)
    }


@app.post("/grupos/zerar-contagem")
def zerar_contagem_grupo(data: ZerarContagemIn, request: Request, db: Session = Depends(get_db)):
    require_admin(request)
    grupo = db.query(Grupo).filter(
        Grupo.id == data.id_grupo, Grupo.id_inventario == data.id_inventario
    ).first()
    if not grupo:
        raise HTTPException(status_code=404, detail="Grupo não encontrado")

    membros = db.query(UsuarioAtivo).filter(
        UsuarioAtivo.id_inventario == data.id_inventario,
        UsuarioAtivo.id_grupo == data.id_grupo
    ).all()

    bipes_apagados = db.query(Bipe).filter(
        Bipe.id_inventario == data.id_inventario,
        Bipe.id_grupo == data.id_grupo
    ).delete(synchronize_session=False)

    for m in membros:
        clear_user_lock_notice(data.id_inventario, m.usuario)

    grupo.status = "RESERVADO" if membros else "DISPONIVEL"
    db.commit()

    return {
        "success": True, "grupo": grupo.nome, "id_grupo": grupo.id,
        "bipes_apagados": int(bipes_apagados or 0),
        "membros_mantidos": [m.usuario for m in membros], "count": 0
    }


@app.post("/grupos/resetar")
def resetar_grupo(data: ResetarGrupoIn, db: Session = Depends(get_db)):
    grupo = db.query(Grupo).filter(
        Grupo.id == data.id_grupo, Grupo.id_inventario == data.id_inventario
    ).first()
    if not grupo:
        raise HTTPException(status_code=404, detail="Grupo não encontrado")

    usuario_no_inventario = obter_grupo_ativo_do_usuario(db, data.usuario, data.id_inventario)
    if not usuario_no_inventario:
        inventario = db.get(Inventario, data.id_inventario)
        if inventario is None:
            raise HTTPException(status_code=404, detail="Inventário não encontrado")

    membros = db.query(UsuarioAtivo).filter(
        UsuarioAtivo.id_inventario == data.id_inventario,
        UsuarioAtivo.id_grupo == data.id_grupo
    ).all()

    for m in membros:
        set_user_lock_notice(data.id_inventario, m.usuario, grupo.nome, "Administrador", "RESETADO")

    bipes_apagados = db.query(Bipe).filter(
        Bipe.id_inventario == data.id_inventario,
        Bipe.id_grupo == data.id_grupo
    ).delete(synchronize_session=False)

    db.query(UsuarioAtivo).filter(
        UsuarioAtivo.id_inventario == data.id_inventario,
        UsuarioAtivo.id_grupo == data.id_grupo
    ).delete(synchronize_session=False)

    grupo.status = "DISPONIVEL"
    db.commit()

    return {
        "success": True, "grupo": grupo.nome, "id_grupo": grupo.id,
        "bipes_apagados": int(bipes_apagados or 0),
        "membros_removidos": [m.usuario for m in membros]
    }


@app.get("/admin/painel")
def admin_painel(db: Session = Depends(get_db)):
    inventarios = db.query(Inventario).all()
    usuarios = db.query(UsuarioAtivo).all()
    grupos = db.query(Grupo).all()
    bipes_raw = db.query(Bipe).all()
    itens_estoque = db.query(Estoque).filter(Estoque.ativo.is_(True)).all()

    estoque_por_ean = {str(normalizar_ean(item.ean or "")): item for item in itens_estoque}

    resumo_grupos = []
    for g in grupos:
        total = contar_bipes_grupo(db, g.id_inventario, g.id)
        membros = db.query(UsuarioAtivo).filter(
            UsuarioAtivo.id_inventario == g.id_inventario,
            UsuarioAtivo.id_grupo == g.id
        ).all()
        resumo_grupos.append({
            "id": g.id, "nome": g.nome, "id_inventario": g.id_inventario,
            "meta": g.meta, "status": g.status, "colaborativo": g.colaborativo,
            "membros": [m.usuario for m in membros], "bipes": total
        })

    # Expande bipes agrupados para exibição item por item no painel
    grupos_map = {str(g.id): g for g in grupos}
    bipes_out = []
    for b in bipes_raw:
        item = estoque_por_ean.get(normalizar_ean(b.ean or ""))
        ref = item.produto if item else ""
        cor = item.cor_produ if item else ""
        tamanho = item.tamanho if item else ""
        ref_cor = item.ref_cor if item else ""
        hora = b.atualizado_em.isoformat() if getattr(b, "atualizado_em", None) else (
            b.criado_em.isoformat() if getattr(b, "criado_em", None) else ""
        )
        for i in range(int(b.quantidade or 1)):
            bipes_out.append({
                "usuario": b.usuario,
                "id_inventario": b.id_inventario,
                "id_grupo": b.id_grupo,
                "grupo_nome": b.grupo_nome,
                "id": int(b.id) * 10000 + i,
                "ean": b.ean,
                "hora": hora,
                "label_compact": f"{ref_cor} {tamanho}".strip() if item else "",
                "ref": ref, "cor": cor, "tamanho": tamanho,
                "grade": tamanho, "filial": "",
                "ref_cor": ref_cor, "nao_encontrado": item is None
            })

    total_estoque = int(sum(int(getattr(i, "quantidade", 0) or 0) for i in itens_estoque))
    grupos_concluidos_ids = {str(g.id) for g in grupos if str(getattr(g, "status", "")) == "CONCLUIDO"}
    total_consolidado_fechado = sum(
        int(b.quantidade or 1)
        for b in bipes_raw
        if str(getattr(b, "id_grupo", "")) in grupos_concluidos_ids
    )
    percentual_consolidado = round((total_consolidado_fechado / max(1, total_estoque)) * 100, 1) if total_estoque else 0.0

    return {
        "success": True,
        "inventarios": [{"id": i.id, "nome": i.nome, "senha": i.senha, "status": i.status} for i in inventarios],
        "usuarios_ativos": [
            {"usuario": u.usuario, "id_inventario": u.id_inventario, "id_grupo": u.id_grupo, "grupo_nome": u.grupo_nome}
            for u in usuarios
        ],
        "grupos": resumo_grupos,
        "bipes": bipes_out,
        "totalEstoque": total_estoque,
        "totalConsolidado": int(total_consolidado_fechado),
        "percentualConsolidado": percentual_consolidado
    }


@app.get("/consolidado")
def get_consolidado(id_inventario: str = "", request: Request = None, db: Session = Depends(get_db)):
    if request is not None:
        require_admin(request)
    rows = listar_consolidado_rows(db, norm_txt(id_inventario))
    return {"success": True, "itens": rows, "total": len(rows)}


@app.delete("/consolidado/{bipe_id}")
def excluir_consolidado(bipe_id: int, request: Request, db: Session = Depends(get_db)):
    require_admin(request)
    # bipe_id expandido = id_real * 10000 + offset
    bipe_id_real = bipe_id // 10000 if bipe_id >= 10000 else bipe_id
    bipe = db.query(Bipe).filter(Bipe.id == bipe_id_real).first()
    if not bipe:
        raise HTTPException(status_code=404, detail="Linha consolidada não encontrada")

    grupo = db.query(Grupo).filter(Grupo.id == bipe.id_grupo).first()
    if not grupo or norm_txt(grupo.status) != "CONCLUIDO":
        raise HTTPException(status_code=400, detail="A linha informada não pertence ao consolidado fechado")

    if bipe.quantidade > 1:
        bipe.quantidade -= 1
        db.commit()
    else:
        db.delete(bipe)
        db.commit()

    total = len(listar_consolidado_rows(db, norm_txt(bipe.id_inventario)))
    return {"success": True, "id": bipe_id, "total": total}


@app.patch("/consolidado/{bipe_id}")
def editar_consolidado(bipe_id: int, data: ConsolidadoUpdateIn, request: Request, db: Session = Depends(get_db)):
    require_admin(request)
    bipe_id_real = bipe_id // 10000 if bipe_id >= 10000 else bipe_id
    bipe = db.query(Bipe).filter(Bipe.id == bipe_id_real).first()
    if not bipe:
        raise HTTPException(status_code=404, detail="Linha consolidada não encontrada")
    grupo = db.query(Grupo).filter(Grupo.id == bipe.id_grupo).first()
    if not grupo or norm_txt(grupo.status) != "CONCLUIDO":
        raise HTTPException(status_code=400, detail="A linha informada não pertence ao consolidado fechado")

    if data.ean is not None:
        novo_ean = normalizar_ean(data.ean)
        if not novo_ean:
            raise HTTPException(status_code=400, detail="EAN inválido")
        bipe.ean = novo_ean

    db.commit()
    db.refresh(bipe)
    item = next(
        (r for r in listar_consolidado_rows(db, norm_txt(bipe.id_inventario)) if int(r["id"]) // 10000 == int(bipe.id)),
        None
    )
    return {"success": True, "item": item}


@app.get("/estoque/confronto")
def confrontar_estoque(id_inventario: str = "", db: Session = Depends(get_db)):
    return montar_confronto_estoque(db, norm_txt(id_inventario))


@app.get("/estoque/confronto/relatorio")
def gerar_relatorio_confronto(id_inventario: str = "", db: Session = Depends(get_db)):
    confronto = montar_confronto_estoque(db, norm_txt(id_inventario))
    if not confronto.get("success"):
        raise HTTPException(status_code=400, detail=confronto.get("message") or "Falha ao gerar relatório")

    wb = Workbook()
    agora = agora_brasil().strftime("%d/%m/%Y %H:%M")
    inv = norm_txt(id_inventario) or "—"

    def montar_sheet(ws, titulo, cabecalho, linhas):
        bloco = [
            ["RELATÓRIO DE CONFRONTO DE ESTOQUE", "", "", "", "", ""],
            ["Inventário:", inv, "Gerado em:", agora, "", ""],
            ["", "", "", "", "", ""],
            [titulo, "", "", "", "", ""],
            cabecalho
        ]
        bloco.extend(linhas if linhas else [["(nenhum)", "", "", "", "", ""]])
        for row in bloco:
            ws.append(row)
        ws["A1"].font = Font(bold=True, size=12)
        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 18

    ws_acima = wb.active
    ws_acima.title = "ACIMA"
    montar_sheet(ws_acima, f"ACIMA DO ESTOQUE ({len(confronto['acima'])} itens)",
        ["Referência", "Cor", "Tamanho", "Qtd Estoque", "Qtd Bipada", "Diferença"],
        [[i.get("ref", ""), i.get("cor", ""), i.get("grade", ""), int(i.get("qtdEstoque", 0)), int(i.get("qtdBipada", 0)), int(i.get("qtdBipada", 0)) - int(i.get("qtdEstoque", 0))] for i in confronto["acima"]])

    for titulo_sheet, chave, cabecalho_extra, extra_fn in [
        ("FALTANDO", "proximo", "% Bipado", lambda i: f"{i.get('pct', 0)}%"),
        ("ABAIXO50", "abaixo50", "% Bipado", lambda i: f"{i.get('pct', 0)}%"),
    ]:
        ws = wb.create_sheet(title=titulo_sheet)
        montar_sheet(ws, f"{titulo_sheet} ({len(confronto[chave])} itens)",
            ["Referência", "Cor", "Tamanho", "Qtd Estoque", "Qtd Bipada", cabecalho_extra],
            [[i.get("ref", ""), i.get("cor", ""), i.get("grade", ""), int(i.get("qtdEstoque", 0)), int(i.get("qtdBipada", 0)), extra_fn(i)] for i in confronto[chave]])

    ws_exato = wb.create_sheet(title="EXATO")
    montar_sheet(ws_exato, f"EXATO ({len(confronto['exato'])} itens)",
        ["Referência", "Cor", "Tamanho", "Qtd Estoque", "Qtd Bipada", "Diferença"],
        [[i.get("ref", ""), i.get("cor", ""), i.get("grade", ""), int(i.get("qtdEstoque", 0)), int(i.get("qtdBipada", 0)), 0] for i in confronto["exato"]])

    ws_ne = wb.create_sheet(title="NAO_ENCONTRADOS")
    montar_sheet(ws_ne, f"NAO ENCONTRADOS ({len(confronto['naoEncontrados'])} itens)",
        ["EAN", "Referência", "Cor", "Tamanho", "Qtd Estoque", "Qtd Bipada"],
        [[i.get("ean", ""), i.get("ref", ""), i.get("cor", ""), i.get("grade", ""), int(i.get("qtdEstoque", 0)), int(i.get("qtdBipada", 0))] for i in confronto["naoEncontrados"]])

    buf = io.BytesIO()
    wb.save(buf)
    conteudo = buf.getvalue()
    nome_base = f"relatorio_confronto_{inv or 'geral'}_{agora_brasil().strftime('%d%m%Y_%H%M%S')}.xlsx"
    nome_seguro = re.sub(r'[^A-Za-z0-9._-]+', '_', nome_base)
    headers = {
        "Content-Disposition": f"attachment; filename=\"{nome_seguro}\"; filename*=UTF-8''{quote(nome_base)}",
        "Access-Control-Expose-Headers": "Content-Disposition",
        "Cache-Control": "no-store"
    }
    return Response(content=conteudo, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)


@app.patch("/inventarios/{id_inventario}/fechar")
def fechar_inventario(id_inventario: str, request: Request, db: Session = Depends(get_db)):
    require_admin(request)
    inv = db.get(Inventario, id_inventario)
    if not inv:
        raise HTTPException(status_code=404, detail="Inventário não encontrado")
    inv.status = "FECHADO"
    db.commit()
    return {"success": True, "message": "Inventário fechado com sucesso.", "id": inv.id, "status": inv.status}


@app.delete("/inventarios/{id_inventario}")
def excluir_inventario(id_inventario: str, request: Request, db: Session = Depends(get_db)):
    require_admin(request)
    inv = db.get(Inventario, id_inventario)
    if not inv:
        raise HTTPException(status_code=404, detail="Inventário não encontrado")

    db.query(EtiquetaPendente).filter(EtiquetaPendente.id_inventario == id_inventario).delete(synchronize_session=False)
    db.query(Bipe).filter(Bipe.id_inventario == id_inventario).delete(synchronize_session=False)
    db.query(UsuarioAtivo).filter(UsuarioAtivo.id_inventario == id_inventario).delete(synchronize_session=False)
    db.query(Grupo).filter(Grupo.id_inventario == id_inventario).delete(synchronize_session=False)
    clear_inventory_lock_notices(id_inventario)
    db.delete(inv)
    db.commit()
    return {"success": True, "message": "Inventário excluído com sucesso.", "id": id_inventario}


@app.delete("/bipes/{bipe_id}")
def excluir_bipe(bipe_id: int, request: Request, db: Session = Depends(get_db)):
    require_admin(request)
    bipe_id_real = bipe_id // 10000 if bipe_id >= 10000 else bipe_id
    bipe = db.query(Bipe).filter(Bipe.id == bipe_id_real).first()
    if not bipe:
        raise HTTPException(status_code=404, detail="Bipe não encontrado")
    id_inventario = bipe.id_inventario
    id_grupo = bipe.id_grupo

    if bipe.quantidade > 1:
        bipe.quantidade -= 1
        bipe.atualizado_em = agora_brasil()
        db.commit()
    else:
        db.delete(bipe)
        db.commit()

    total_grupo = contar_bipes_grupo(db, id_inventario, id_grupo)
    return {"success": True, "id": bipe_id, "newCount": int(total_grupo)}


@app.get("/etiquetas")
def get_etiquetas(id_inventario: str = "", request: Request = None, db: Session = Depends(get_db)):
    if request is not None:
        require_admin(request)
    return {"success": True, "etiquetas": listar_etiquetas_pendentes(db, id_inventario)}


@app.delete("/etiquetas")
def limpar_etiquetas(id_inventario: str = "", request: Request = None, db: Session = Depends(get_db)):
    if request is not None:
        require_admin(request)
    query = db.query(EtiquetaPendente)
    if norm_txt(id_inventario):
        query = query.filter(EtiquetaPendente.id_inventario == norm_txt(id_inventario))
    removidas = query.delete(synchronize_session=False)
    db.commit()
    return {"success": True, "removidas": int(removidas or 0)}


@app.get("/etiquetas/pdf")
def etiquetas_pdf(id_inventario: str = "", request: Request = None, db: Session = Depends(get_db)):
    if request is not None:
        require_admin(request)
    etiquetas = listar_etiquetas_pendentes(db, id_inventario)
    conteudo = gerar_pdf_etiquetas_bytes(etiquetas)
    nome = f"etiquetas_{norm_txt(id_inventario) or 'geral'}.pdf"
    headers = {
        "Content-Disposition": f'attachment; filename="{nome}"',
        "Access-Control-Expose-Headers": "Content-Disposition",
        "Cache-Control": "no-store"
    }
    return Response(content=conteudo, media_type="application/pdf", headers=headers)


@app.get("/estoque/validar")
def validar_estoque_por_ean(ean: str = "", db: Session = Depends(get_db)):
    ean_norm = normalizar_ean(ean)
    if not ean_norm:
        return {"success": True, "encontrado": False, "ean": "", "item": None, "info": None}
    item = buscar_item_estoque_por_ean(db, ean_norm)
    if not item:
        return {"success": True, "encontrado": False, "ean": ean_norm, "item": None, "info": None}
    ref = norm_txt(getattr(item, "produto", ""))
    cor = norm_txt(getattr(item, "cor_produ", ""))
    tamanho = norm_txt(getattr(item, "tamanho", ""))
    ref_cor = norm_txt(getattr(item, "ref_cor", "")) or montar_ref_cor(ref, cor)
    info = {
        "ref": ref, "cor": cor, "tamanho": tamanho, "grade": tamanho,
        "refCor": ref_cor, "labelCompact": f"{ref_cor} {tamanho}".strip(),
        "label": f"{ref_cor} {tamanho}".strip()
    }
    return {"success": True, "encontrado": True, "ean": ean_norm, "item": info, "info": info}


@app.get("/estoque/mapa-mini")
def get_mapa_estoque_mini(id_inventario: str = "", db: Session = Depends(get_db)):
    itens = db.query(Estoque).filter(Estoque.ativo.is_(True)).all()
    mapa = {}
    for item in itens:
        ean = normalizar_ean(getattr(item, "ean", ""))
        if not ean:
            continue
        ref = norm_txt(getattr(item, "produto", ""))
        cor = norm_txt(getattr(item, "cor_produ", ""))
        tamanho = norm_txt(getattr(item, "tamanho", ""))
        ref_cor = norm_txt(getattr(item, "ref_cor", "")) or montar_ref_cor(ref, cor)
        texto = f"{ref_cor} {tamanho}".strip()
        if texto:
            mapa[ean] = texto
    return {
        "success": True, "idInventario": norm_txt(id_inventario),
        "total": len(mapa), "mapa": mapa,
        "geradoEm": agora_brasil().isoformat()
    }


@app.get("/consolidado/exportar-eans")
def exportar_eans_consolidado(id_inventario: str = Query(""), request: Request = None, db: Session = Depends(get_db)):
    if request is not None:
        require_admin(request)
    rows = listar_consolidado_rows(db, norm_txt(id_inventario))

    wb = Workbook()
    ws = wb.active
    ws.title = "EANS"
    ws["A1"] = "EAN"
    ws["A1"].font = Font(bold=True)

    row_num = 2
    for r in rows:
        ean = norm_txt(r.get("ean", ""))
        if ean:
            ws.cell(row=row_num, column=1, value=ean)
            row_num += 1

    ws.column_dimensions["A"].width = 22
    buffer = io.BytesIO()
    wb.save(buffer)
    conteudo = buffer.getvalue()
    nome_base = f"eans_consolidado_{norm_txt(id_inventario) or 'geral'}_{agora_brasil().strftime('%d%m%Y_%H%M%S')}.xlsx"
    nome_seguro = re.sub(r'[^A-Za-z0-9._-]+', '_', nome_base)

    return Response(
        content=conteudo,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nome_seguro}"'}
    )


# ── IMPORTAR ESTOQUE (xlsx/csv) ───────────────────────────────────────────────
@app.post("/estoque/importar")
async def importar_estoque(
    arquivo: UploadFile = File(...),
    substituir_tudo: bool = Query(True),
    db: Session = Depends(get_db)
):
    nome = (arquivo.filename or "").lower()
    conteudo = await arquivo.read()
    linhas = []

    try:
        if nome.endswith(".xlsx"):
            wb = load_workbook(filename=io.BytesIO(conteudo), read_only=True, data_only=True)
            ws = wb.active
            headers = [str(c.value or "").strip().upper() for c in next(ws.iter_rows(min_row=1, max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                linhas.append({headers[i]: str(v or "").strip() for i, v in enumerate(row) if i < len(headers)})
        elif nome.endswith(".csv"):
            texto = conteudo.decode("utf-8-sig", errors="replace")
            reader = csv.DictReader(io.StringIO(texto))
            for row in reader:
                linhas.append({k.strip().upper(): str(v or "").strip() for k, v in row.items()})
        else:
            raise HTTPException(status_code=400, detail="Formato não suportado. Use .xlsx ou .csv")
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Erro ao ler arquivo: {e}")

    if substituir_tudo:
        db.query(Estoque).update({"ativo": False}, synchronize_session=False)
        db.commit()

    inseridos, ignorados = 0, 0
    for linha in linhas:
        # Aceita tanto CODIGO_BARR quanto CODIGO_BARRA
        ean_raw = linha.get("CODIGO_BARR") or linha.get("CODIGO_BARRA") or linha.get("EAN") or ""
        ean = normalizar_ean(ean_raw)
        if not ean:
            ignorados += 1
            continue

        produto = linha.get("PRODUTO", "")
        cor = linha.get("COR_PRODU", "")
        tamanho = linha.get("TAMANHO", "")
        quantidade_raw = linha.get("QUANTIDADE", "0")
        try:
            quantidade = int(float(quantidade_raw or "0"))
        except (ValueError, TypeError):
            quantidade = 0

        ref_cor = f"{produto}{cor}".strip()

        existente = db.query(Estoque).filter(Estoque.ean == ean).first()
        if existente:
            existente.produto = produto
            existente.cor_produ = cor
            existente.tamanho = tamanho
            existente.quantidade = quantidade
            existente.ref_cor = ref_cor
            existente.ativo = True
        else:
            db.add(Estoque(
                produto=produto, cor_produ=cor, tamanho=tamanho,
                quantidade=quantidade, ean=ean, ref_cor=ref_cor, ativo=True
            ))
        inseridos += 1

    db.commit()
    total_estoque = db.query(func.count(Estoque.id)).filter(Estoque.ativo.is_(True)).scalar() or 0
    return {"success": True, "inseridos": inseridos, "ignorados": ignorados, "total_estoque": int(total_estoque)}
